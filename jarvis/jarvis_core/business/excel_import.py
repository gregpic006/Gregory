"""Lecture des rapports Excel.

Un back-office Windows exporte rarement en CSV par defaut: il propose Excel,
et c'est ce que l'utilisateur obtient sans y penser. Refuser le .xlsx, c'est
lui demander de refaire son export — une manipulation de plus, pour une raison
qui ne le concerne pas.

Ce module ne fait qu'une chose: transformer la premiere feuille en texte
delimite, puis laisser le lecteur CSV faire son travail habituel. Toute la
logique d'interpretation — lignes de titre a ignorer, formats de date,
montants a la francaise, colonnes reconnues — reste au meme endroit. Un second
analyseur en parallele finirait par diverger du premier.

Trois precautions.

**Lecture seule et sans formules.** `data_only=True` rend la derniere valeur
calculee, jamais la formule. Une feuille de rapport ne doit pas etre evaluee
par nous.

**Borne en taille.** Un classeur de plusieurs dizaines de milliers de lignes
n'est pas un rapport quotidien; on s'arrete avant de le charger entierement.

**Les valeurs gardent leur forme.** Les dates Excel arrivent en `datetime`; on
les rend en ISO plutot qu'en texte localise, parce que c'est ce que le lecteur
CSV sait reconnaitre a coup sur.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

#: Au-dela, ce n'est pas un rapport de caisse.
MAX_ROWS = 20_000
SUFFIXES = frozenset({".xlsx", ".xlsm"})


def is_excel(name: str) -> bool:
    """Vrai si ce nom de fichier designe un classeur lisible ici.

    `.xls` (ancien format binaire) n'en fait pas partie: le lire demanderait
    une dependance de plus, et le dire franchement vaut mieux que d'echouer
    avec un message obscur.
    """
    lowered = name.lower()
    return any(lowered.endswith(suffix) for suffix in SUFFIXES)


def _cell(value: Any) -> str:
    """Rend une cellule sous une forme que le lecteur CSV comprend."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # `2410.0` doit rester « 2410 »: un separateur decimal parasite
        # ferait echouer la lecture du montant.
        return str(int(value)) if value.is_integer() else repr(value)
    return str(value)


def excel_to_csv(source: Path | bytes, *, max_rows: int = MAX_ROWS) -> str:
    """Convertit la premiere feuille en texte CSV.

    Args:
        source: chemin du classeur, ou son contenu (piece jointe de courriel).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depend de l'installation
        raise ExcelUnavailable(
            "Le lecteur de fichiers Excel n'est pas installe. Relance JARVIS: "
            "il l'installe tout seul au demarrage."
        ) from exc

    handle: Any = io.BytesIO(source) if isinstance(source, bytes) else source
    try:
        workbook = load_workbook(handle, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl leve des types varies
        raise ExcelUnreadable(
            "Ce fichier Excel n'a pas pu etre ouvert. S'il vient d'un export, "
            "demande plutot un CSV."
        ) from exc

    try:
        sheet = workbook.worksheets[0]
        buffer = io.StringIO()
        writer = csv.writer(buffer, lineterminator="\n")
        for written, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Les lignes entierement vides separent souvent les sections d'un
            # rapport; les garder ne nuit pas, le lecteur CSV les ignore.
            writer.writerow([_cell(value) for value in row])
            if written >= max_rows:
                logger.warning("classeur tronque a %s lignes", max_rows)
                break
        return buffer.getvalue()
    finally:
        workbook.close()


class ExcelUnavailable(RuntimeError):
    """Le lecteur Excel n'est pas installe."""


class ExcelUnreadable(ValueError):
    """Le classeur existe mais ne s'ouvre pas."""
